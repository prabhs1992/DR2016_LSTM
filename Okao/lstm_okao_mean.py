from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import time

import tensorflow.python.platform

import numpy as np
import tensorflow as tf

from tensorflow.models.rnn import rnn_cell
from tensorflow.models.rnn import seq2seq
from tensorflow.models.rnn.ptb import reader
from openpyxl import load_workbook
import math
import os
import re
from collections import defaultdict
import random

flags = tf.flags
logging = tf.logging

flags.DEFINE_string(
    "model", "small",
    "A type of model. Possible options are: small, medium, large.")
flags.DEFINE_string("data_path", None, "data_path")

FLAGS = flags.FLAGS


class OKAOModel(object):

  def __init__(self, is_training, config):
    self.batch_size = batch_size = config.batch_size
    self.num_steps = num_steps = config.num_steps
    size = config.hidden_size
    output_size = config.output_size

    self._input_data = tf.placeholder(tf.float32, [batch_size, num_steps, size])
    self._targets = tf.placeholder(tf.int32, [batch_size, num_steps])

    lstm_cell = rnn_cell.BasicLSTMCell(size, forget_bias=0.0)
    if is_training and config.keep_prob < 1:
      lstm_cell = rnn_cell.DropoutWrapper(
          lstm_cell, output_keep_prob=config.keep_prob)
    cell = rnn_cell.MultiRNNCell([lstm_cell] * config.num_layers)

    self._initial_state = cell.zero_state(batch_size, tf.float32)

    inputs = self._input_data

    outputs = []
    states = []
    state = self._initial_state
    with tf.variable_scope("RNN"):
      for time_step in range(num_steps):
        if time_step > 0: tf.get_variable_scope().reuse_variables()
        (cell_output, state) = cell(inputs[:, time_step, :], state)
        outputs.append(cell_output)
        states.append(state)

    output = tf.reshape(tf.concat(1, outputs), [-1, size])
    logits = tf.nn.xw_plus_b(output,
                             tf.get_variable("softmax_w", [size, output_size]),
                             tf.get_variable("softmax_b", [output_size]))
    loss = seq2seq.sequence_loss_by_example([logits],
                                            [tf.reshape(self._targets, [-1])],
                                            [tf.ones([batch_size * num_steps])],
                                            output_size)
    self._cost = cost = tf.reduce_sum(loss) / batch_size
    self._final_state = states[-1]
    self._output = output
    self._logits = logits

    if not is_training:
      return

    self._lr = tf.Variable(0.0, trainable=False)
    tvars = tf.trainable_variables()
    grads, _ = tf.clip_by_global_norm(tf.gradients(cost, tvars),
                                      config.max_grad_norm)
    optimizer = tf.train.GradientDescentOptimizer(self.lr)
    self._train_op = optimizer.apply_gradients(zip(grads, tvars))

  def assign_lr(self, session, lr_value):
    session.run(tf.assign(self.lr, lr_value))

  @property
  def input_data(self):
    return self._input_data

  @property
  def targets(self):
    return self._targets

  @property
  def initial_state(self):
    return self._initial_state

  @property
  def cost(self):
    return self._cost

  @property
  def output(self):
    return self._output

  @property
  def logits(self):
    return self._logits

  @property
  def final_state(self):
    return self._final_state

  @property
  def lr(self):
    return self._lr

  @property
  def train_op(self):
    return self._train_op


class SmallConfig(object):
  """Small config."""
  init_scale = 0.1
  learning_rate = 1.0
  max_grad_norm = 5
  num_layers = 1
  num_steps = 20
  hidden_size = 165
  max_epoch = 4
  max_max_epoch = 13
  keep_prob = 1.0
  lr_decay = 0.5
  batch_size = 20
  output_size = 2


class MediumConfig(object):
  """Medium config."""
  init_scale = 0.05
  learning_rate = 1.0
  max_grad_norm = 5
  num_layers = 2
  num_steps = 20
  hidden_size = 165
  max_epoch = 6
  max_max_epoch = 39
  keep_prob = 0.5
  lr_decay = 0.8
  batch_size = 20
  output_size = 2


class LargeConfig(object):
  """Large config."""
  init_scale = 0.04
  learning_rate = 1.0
  max_grad_norm = 10
  num_layers = 2
  num_steps = 20
  hidden_size = 165
  max_epoch = 14
  max_max_epoch = 55
  keep_prob = 0.35
  lr_decay = 1 / 1.15
  batch_size = 20
  output_size = 2


def run_epoch(bool_test, session, m, data, label, eval_op, indices, speaker, verbose=False):
  """Runs the model on the given data."""
  #epoch_size = ((len(data) // m.batch_size) - 1) // m.num_steps
  epoch_size = ((len(data) // m.batch_size)) // m.num_steps
  start_time = time.time()
  costs = 0.0
  iters = 0
  state = m.initial_state.eval()

  ip_data = []
  ip_label = []

  for i in range(((len(data) // m.batch_size)) // m.num_steps):
    temp3d = []
    templ3d = []
    for j in range(m.num_steps):
        temp = data[i*(m.batch_size*m.num_steps)+j*(m.batch_size):i*(m.batch_size*m.num_steps)+(j+1)*m.batch_size,:]
        temp = temp[np.newaxis,:,:]
        templ = label[i*(m.batch_size*m.num_steps)+j*(m.batch_size):i*(m.batch_size*m.num_steps)+(j+1)*m.batch_size]
        templ = np.transpose(templ)
        if len(temp3d) == 0:
            temp3d = temp
            templ3d = templ
        else:
            temp3d = np.concatenate((temp3d,temp),axis=0)
            templ3d = np.concatenate((templ3d,templ),axis=0)
    temp3d = temp3d[np.newaxis,:,:,:]
    templ3d = templ3d[np.newaxis,:,:]
    if len(ip_data) == 0:
        ip_data = temp3d
        ip_label = templ3d
    else:
        ip_data = np.concatenate((ip_data,temp3d),axis=0)
        ip_label = np.concatenate((ip_label,templ3d),axis=0)

  all_logits = []
  for step in range(epoch_size):
    x = ip_data[step]
    if not bool_test:
        y = ip_label[step]
        cost, state, _, output, logits = session.run([m.cost, m.final_state, eval_op, m.output, m.logits],
                                                     {m.input_data: x,
                                                      m.targets: y,
                                                      m.initial_state: state})
        costs += cost
    else:
        #in case of testing don't use labels.
        state, _, output, logits = session.run([m.final_state, eval_op, m.output, m.logits],
                                                     {m.input_data: x,
                                                      m.initial_state: state})
    iters += m.num_steps
    if bool_test == 1:
      if len(all_logits) == 0:
        all_logits = logits
      else:
        all_logits = np.concatenate((all_logits,logits),axis = 0)
    if verbose and step % (epoch_size // 10) == 10:
      print("%.3f perplexity: %.3f speed: %.0f wps" %
            (step * 1.0 / epoch_size, np.exp(costs / iters),
             iters * m.batch_size / (time.time() - start_time)))

  if bool_test == 0:
    return np.exp(costs / iters)
  else:
    exp_logits = np.exp(all_logits)
    sum_exp = np.sum(exp_logits, axis=1)
    # prob = exp_logits/sum_exp
    max = np.argmax(exp_logits, axis=1)
    # logical_arr = max == ip_label #takes a lot of time so removing this
    correct_pred = 0
    f_correct_pred = 0
    # code for framewise prediction:
    # for i in range(len(max)):
    #    if max[i] == ip_label[i]:
    #        correct_pred += 1
    # return (100*correct_pred) / len(ip_label)
    # code for framewise prediction:
    for i in range(len(max)):
        if max[i] == ip_label[i]:
            f_correct_pred += 1
    f_acc = (100 * f_correct_pred) / len(ip_label)
    print("Framewise Test accuracy: %.3f" % f_acc)

    # speaker wise accuracy prediction:
    cntspeakers = 0
    correctsepaker_pred = 0
    for i in range(len(indices)):
        s11cnt = s11pred = 0
        s12cnt = s12pred = 0
        s21cnt = s21pred = 0
        s22cnt = s22pred = 0
        for j in range(indices[i]):
            if speaker[cntspeakers + j] == "S11":
                s11cnt += 1
                if max[cntspeakers + j] == ip_label[cntspeakers + j]:
                    s11pred += 1
            elif speaker[cntspeakers + j] == "S12":
                s12cnt += 1
                if max[cntspeakers + j] == ip_label[cntspeakers + j]:
                    s12pred += 1
            elif speaker[cntspeakers + j] == "S21":
                s21cnt += 1
                if max[cntspeakers + j] == ip_label[cntspeakers + j]:
                    s21pred += 1
            else:
                s22cnt += 1
                if max[cntspeakers + j] == ip_label[cntspeakers + j]:
                    s22pred += 1
        if s11pred > s11cnt / 2:
            correctsepaker_pred += 1
        if s12pred > s12cnt / 2:
            correctsepaker_pred += 1
        if s21pred > s21cnt / 2:
            correctsepaker_pred += 1
        if s22pred > s22cnt / 2:
            correctsepaker_pred += 1
        cntspeakers += indices[i]

    s_acc = (100 * correctsepaker_pred) / (4 * len(indices))
    print("Speakerwise Test accuracy: %.3f" % s_acc)

    # teamwise accuracy prediction:
    cntindices = 0
    for i in range(len(indices)):
        correct_1 = 0
        correct_0 = 0
        total = []
        total.append(0)
        total.append(0)
        for j in range(indices[i]):
            total[ip_label[cntindices + j]] += 1
            if max[cntindices + j] == ip_label[cntindices + j]:
                if ip_label[cntindices + j] == 1:
                    correct_1 += 1
                else:
                    correct_0 += 1
        cntindices += indices[i]
        if correct_1 > total[1] / 2:
            correct_pred += 1
        if correct_0 > total[0] / 2:
            correct_pred += 1
    return (100 * correct_pred) / (2 * len(indices))


def get_config():
  if FLAGS.model == "small":
    return SmallConfig()
  elif FLAGS.model == "medium":
    return MediumConfig()
  elif FLAGS.model == "large":
    return LargeConfig()
  else:
    raise ValueError("Invalid model: %s", FLAGS.model)


def main(unused_args):
    wb = load_workbook('/home/prabhanjan/Downloads/Matlab_MetaData.xlsx', read_only=True)
    dir = '/media/prabhanjan/New/DR/Okao Features/'
    testcnt = 5
    limit = 36

    ws = wb.get_sheet_by_name('Sheet1')

    x10 = np.array([r[10].value for r in ws.iter_rows()])
    x2 = np.array([r[2].value for r in ws.iter_rows()])

    Y = {}
    p2 = re.compile('[A-Z]*([0-9]*_.*)\'')
    for i in range(3,len(x10)):
        m2 = p2.search(x2[i])
        Y[m2.group(1)] = x10[i]
    f_grp = defaultdict(list)
    for filename in os.listdir(dir):
        if not filename.startswith('.'):
            video = re.compile('([0-9]*)_Frontal')
            m = video.search(filename)
            if m != None:
                f_grp[m.group(1)].append(filename)

    cntr = 0
    train_data = []
    train_label = []
    valid_data = []
    validcnt = 0
    valid_label = []
    test_data = []
    test_label = []
    indices = []
    valid_indices = []
    sp = []
    spvalid = []

    #shuffle list before splitting as train/test
    keys = f_grp.keys()
    random.shuffle(keys)

    for key in keys:
        grp = f_grp[key]
        cntr += 1
        if cntr > limit:
            break

        data = []
        label = []
        for index, f in enumerate(grp):
            raw_dt = np.genfromtxt(dir+f, delimiter=',')
            if len(data) == 0:
                temp = raw_dt[0:len(raw_dt[:,1]),1:len(raw_dt[1,:])]
                temp2 = []
                #modification for mean of every 5 frames
                for i in range(0,int(math.floor(len(temp[:,1])/5))):
                    if len(temp2) == 0:
                        temp2 = np.mean(temp[i*5:(i+1)*5-1,:],axis=0)
                        temp2 = temp2.reshape(1,-1)
                    else:
                        tempmean = np.mean(temp[i*5:(i+1)*5-1,:],axis=0).reshape(1,-1)
                        temp2 = np.concatenate((temp2,tempmean))
                if len(temp[:,1])%5 != 0:
                    tempmean = np.mean(temp[math.floor(len(temp[:,1])/5)*5:,:],axis=0).reshape(1,-1)
                    temp2 = np.concatenate((temp2,tempmean))
                data = temp2

            else:
                temp = raw_dt[0:len(raw_dt[:,1]),1:len(raw_dt[1,:])]
                temp2 = []
                #modification for mean of every 5 frames
                for i in range(0,int(math.floor(len(temp[:,1])/5))):
                    if len(temp2) == 0:
                        temp2 = np.mean(temp[i*5:(i+1)*5-1,:],axis=0)
                        temp2 = temp2.reshape(1,-1)
                    else:
                        tempmean = np.mean(temp[i*5:(i+1)*5-1,:],axis=0).reshape(1,-1)
                        temp2 = np.concatenate((temp2,tempmean))
                if len(temp[:,1])%5 != 0:
                    tempmean = np.mean(temp[math.floor(len(temp[:,1])/5)*5:,:],axis=0).reshape(1,-1)
                    temp2 = np.concatenate((temp2,tempmean))
                data = np.concatenate((data, temp2),axis=0)

            pat = re.compile('(S[0-9]*_[A-Z]*)')
            mat = pat.search(f)

            if cntr > limit-(validcnt+testcnt):
                if "S11" in f:
                    extra = 0 if (len(raw_dt[:,1]))%5 == 0 else 1
                    for k in range(int(math.floor((len(raw_dt[:,1]))/5)+extra)):
                        sp.append("S11")
                elif "S12" in f:
                    extra = 0 if (len(raw_dt[:,1]))%5 == 0 else 1
                    for k in range(int(math.floor((len(raw_dt[:,1]))/5)+extra)):
                        sp.append("S12")
                elif "S21" in f:
                    extra = 0 if (len(raw_dt[:,1]))%5 == 0 else 1
                    for k in range(int(math.floor((len(raw_dt[:,1]))/5)+extra)):
                        sp.append("S21")
                else:
                    extra = 0 if (len(raw_dt[:,1]))%5 == 0 else 1
                    for k in range(int(math.floor((len(raw_dt[:,1]))/5)+extra)):
                        sp.append("S22")


            if len(label) == 0:
                if Y[key+"_"+mat.group(0)] == 0:
                    #label = np.tile([0],(len(raw_dt[:,1])-1,1))
                    extra = 0 if (len(raw_dt[:,1]))%5 == 0 else 1
                    label = np.tile([0],(math.floor((len(raw_dt[:,1]))/5)+extra,1))
                else:
                    #label = np.tile([1],(len(raw_dt[:,1])-1,1))
                    extra = 0 if (len(raw_dt[:,1]))%5 == 0 else 1
                    label = np.tile([1],(math.floor((len(raw_dt[:,1]))/5)+extra,1))
            else:
                if Y[key+"_"+mat.group(0)] == 0:
                    #label = np.concatenate((label,np.tile([0],(len(raw_dt[:,1])-1,1))),axis=0)
                    extra = 0 if (len(raw_dt[:,1]))%5 == 0 else 1
                    label = np.concatenate((label,np.tile([0],(math.floor((len(raw_dt[:,1]))/5)+extra,1))),axis=0)
                else:
                    #label = np.concatenate((label,np.tile([1],(len(raw_dt[:,1])-1,1))),axis=0)
                    extra = 0 if (len(raw_dt[:,1]))%5 == 0 else 1
                    label = np.concatenate((label,np.tile([1],(math.floor((len(raw_dt[:,1]))/5)+extra,1))),axis=0)

        #shuffle_seed = np.random.permutation(len(label))
        #label = label[shuffle_seed]
        #data = data[shuffle_seed]
        if cntr <= limit-(validcnt+testcnt):
            if len(train_data) == 0:
                train_data = data
                train_label = label
            else:
                train_data = np.concatenate((train_data,data),axis = 0)
                train_label = np.concatenate((train_label,label),axis = 0)
        elif cntr <= limit - (testcnt):
                spvalid = sp
                if cntr == limit - (testcnt):
                    sp = []
                if len(valid_data) == 0:
                    valid_data = data
                    valid_label = label
                    valid_indices = [len(label)]
                else:
                    valid_data = np.concatenate((valid_data, data), axis=0)
                    valid_label = np.concatenate((valid_label, label), axis=0)
                    valid_indices = np.concatenate((valid_indices, [len(label)],), axis=0)
        else:
            if len(test_data) == 0:
                test_data = data
                test_label = label
                indices = [len(label)]
            else:
                test_data = np.concatenate((test_data, data), axis=0)
                test_label = np.concatenate((test_label, label), axis=0)
                indices = np.concatenate((indices, [len(label)],), axis=0)

    config = get_config()
    eval_config = get_config()
    eval_config.batch_size = 1
    eval_config.num_steps = 1

    with tf.Graph().as_default(), tf.Session() as session:
        initializer = tf.random_uniform_initializer(-config.init_scale,
                                                    config.init_scale)
        with tf.variable_scope("model", reuse=None, initializer=initializer):
          m = OKAOModel(is_training=True, config=config)
        with tf.variable_scope("model", reuse=True, initializer=initializer):
          #mvalid = OKAOModel(is_training=False, config=config)
          mtest = OKAOModel(is_training=False, config=eval_config)

        tf.initialize_all_variables().run()

        for i in range(config.max_max_epoch):
          lr_decay = config.lr_decay ** max(i - config.max_epoch, 0.0)
          m.assign_lr(session, config.learning_rate * lr_decay)

          print("Epoch: %d Learning rate: %.3f" % (i + 1, session.run(m.lr)))
          train_perplexity = run_epoch(0, session, m, train_data, train_label, m.train_op,
                                       [], [], verbose=True)
          print("Epoch: %d Train Perplexity: %.3f" % (i + 1, train_perplexity))
          #valid_perplexity = run_epoch(session, mvalid, valid_data, tf.no_op())
          #print("Epoch: %d Valid Perplexity: %.3f" % (i + 1, valid_perplexity))
          #print(session.run(m._final_state))


        test_accuracy = run_epoch(1, session, mtest, test_data, test_label, tf.no_op(), indices, sp)
        print("Teamwise Test accuracy: %.3f" % test_accuracy)


if __name__ == "__main__":
  tf.app.run()
