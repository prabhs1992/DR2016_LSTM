from __future__ import absolute_import
from __future__ import division
from __future__ import print_function

import numpy as np
from openpyxl import load_workbook
import pandas as pd
import re
from sklearn import tree
from sklearn import metrics

wb = load_workbook('/home/prabhanjan/Downloads/Matlab_MetaData.xlsx', read_only=True)
liwc_data = '/media/prabhanjan/New/DR/LIWC Features/LIWC_Data.xlsx'
limit = 288

testcnt = 0


ws = wb.get_sheet_by_name('Sheet1')

x10 = np.array([r[10].value for r in ws.iter_rows()])
x2 = np.array([r[2].value for r in ws.iter_rows()])

Y = {}
p2 = re.compile('[A-Z]*([0-9]*_.*)\'')
for i in range(3, len(x10)):
    m2 = p2.search(x2[i])
    Y[m2.group(1)] = x10[i]

raw_dt = pd.read_excel(liwc_data,'Sheet1')

data = raw_dt.ix[:,2:raw_dt.shape[1]]
np_data = data.as_matrix()
label = []

for i in range(len(raw_dt)):
    f = raw_dt.ix[i,0]
    f = f.encode('ascii', 'ignore')
    newf = f[1:len(f)-5]
    newf = newf if (newf[0] is not 'N') else (newf[1:len(newf)])
    key = (newf+'LS') if (newf[len(newf)-1] == 'C') else (newf+'SS')
    if len(label) == 0:
        if Y[key] == 0:
            label = np.tile([0], (1, 1))
        else:
            label = np.tile([1], (1, 1))
    else:
        if Y[key] == 0:
            label = np.concatenate((label, np.tile([0], (1, 1))), axis=0)
        else:
            label = np.concatenate((label, np.tile([1], (1, 1))), axis=0)

#shuffle data sets :
p = np.random.permutation(len(label))
label = label[p,:]
np_data = np_data[p,:]

train_data = np_data[0:limit-testcnt,:]
train_label = label[0:limit-testcnt]
test_data = np_data[limit-testcnt:limit,:]
test_label = label[limit-testcnt:limit]

clf = tree.DecisionTreeClassifier()

total_acc = 0
#prediction = clf.predict(test_data)
for i in range(limit):
    tr_data = np.delete(train_data,i,0)
    tr_label = np.delete(train_label,i,0)
    clf = clf.fit(tr_data, tr_label)
    pred = clf.predict(train_data[i,:])
    acc = 100*metrics.accuracy_score(train_label[i,:],pred)
    total_acc += acc

print('LOOCV accuracy = %.3f' %(total_acc/limit))
